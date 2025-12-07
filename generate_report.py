import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_test_report_xlsx(
    project_name: str,
    scenario_name: str,
    results_path: str = None,
    api_config_path: str = "configs/api_config_new.json",
    output_path: str = None,
):
    """
    Generate a styled Excel report for a given project + scenario
    using results.json and api_config_new.json.
    """

    # ----- Resolve default paths -----
    if results_path is None:
        # default: results/<project_name>/results.json
        results_path = os.path.join("results", "results.json")

    if output_path is None:
        os.makedirs("reports", exist_ok=True)
        
        timestamp = datetime.now().strftime("%d%m%y%S")  # DDMMYYSS format
        
        output_filename = f"{timestamp}_{project_name}_{scenario_name}_report.xlsx"
        output_filename = output_filename.replace(" ", "_")  # remove spaces
        
        output_path = os.path.join("reports", output_filename)

    # ----- Load JSON data -----
    if not os.path.exists(results_path):
        raise FileNotFoundError(f"results.json not found at: {results_path}")

    with open(results_path, "r", encoding="utf-8") as f:
        results = json.load(f)

    scenario_data = results.get(scenario_name)
    if scenario_data is None:
        raise ValueError(f"Scenario '{scenario_name}' not found in results.json")

    if not os.path.exists(api_config_path):
        raise FileNotFoundError(f"api_config_new.json not found at: {api_config_path}")

    with open(api_config_path, "r", encoding="utf-8") as f:
        api_config = json.load(f)

    # ----- Compute metrics -----
    total_apis = len(scenario_data)

    # testcases per API (count entries per API)
    tests_per_api = {api_name: len(tc_dict) for api_name, tc_dict in scenario_data.items()}
    max_testcases_per_api = max(tests_per_api.values()) if tests_per_api else 0

    total_passed = 0
    total_failed = 0

    # Error code → count (for failed only)
    error_code_counts = {}  # status_code -> count

    # (endpoint, error_code) -> [test_indices]
    failed_endpoints = {}   # (endpoint, code) -> list of testcase indices

    for api_name, testcases in scenario_data.items():
        for idx_str, res in testcases.items():
            try:
                test_index = int(idx_str)
            except ValueError:
                # in case keys aren't numeric strings
                test_index = idx_str

            status = res.get("status_code")
            error = res.get("error")

            # Normalize status_code to int if possible
            code_int = None
            if isinstance(status, int):
                code_int = status
            elif isinstance(status, str):
                try:
                    code_int = int(status)
                except ValueError:
                    code_int = None

            # Decide pass/fail: 2xx + no error -> passed, else failed
            if code_int is not None and 200 <= code_int < 300 and (error is None or error == ""):
                total_passed += 1
            else:
                total_failed += 1

                if code_int is not None:
                    error_code_counts[code_int] = error_code_counts.get(code_int, 0) + 1

                    # Build endpoint from api_config
                    cfg = api_config.get(api_name, {})
                    base_url = cfg.get("url", "")
                    path = cfg.get("path", "")
                    endpoint = f"{base_url}{path}" if (base_url or path) else api_name

                    key = (endpoint, code_int)
                    failed_endpoints.setdefault(key, []).append(str(test_index))

    # ----- Create Excel workbook -----
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"

    # Styles
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="4F81BD")  # blue header
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="D9D9D9")  # light grey for section titles
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    row = 1

    # ----- Header: Project & Scenario -----
    ws[f"A{row}"] = "Project name:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = project_name
    row += 1

    ws[f"A{row}"] = "Scenario name:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = scenario_name
    row += 2  # blank line after header

    # ----- Summary numbers -----
    ws[f"A{row}"] = "Total no. of APIs:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = total_apis
    row += 1

    ws[f"A{row}"] = "Total no. of testcases per API:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = max_testcases_per_api
    row += 1

    ws[f"A{row}"] = "Total no. of testcases passed:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = total_passed
    row += 1

    ws[f"A{row}"] = "Total no. of testcases failed:"
    ws[f"A{row}"].font = bold_font
    ws[f"B{row}"] = total_failed
    row += 2

    # ----- Error Code Summary Table -----
    ws[f"A{row}"] = "Error Code Summary"
    ws[f"A{row}"].font = bold_font
    ws[f"A{row}"].fill = section_fill
    row += 1

    # Table headers
    ws[f"A{row}"] = "Error code"
    ws[f"B{row}"] = "Count"
    for col in ("A", "B"):
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].alignment = center
    row += 1

    # Table rows
    if error_code_counts:
        for code, count in sorted(error_code_counts.items()):
            ws[f"A{row}"] = code
            ws[f"A{row}"].alignment = center
            ws[f"B{row}"] = count
            ws[f"B{row}"].alignment = center
            row += 1
    else:
        ws[f"A{row}"] = "No failed testcases."
        row += 1

    row += 1  # blank line

    # ----- Failed API endpoints Table -----
    ws[f"A{row}"] = "Failed API Endpoints"
    ws[f"A{row}"].font = bold_font
    ws[f"A{row}"].fill = section_fill
    row += 1

    # Table headers
    ws[f"A{row}"] = "Failed API endpoint"
    ws[f"B{row}"] = "Error code"
    ws[f"C{row}"] = "Test case indices"
    for col in ("A", "B", "C"):
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].alignment = center
    row += 1

    if failed_endpoints:
        for (endpoint, code), indices in sorted(failed_endpoints.items(), key=lambda x: (x[0][0], x[0][1])):
            ws[f"A{row}"] = endpoint
            ws[f"A{row}"].alignment = left

            ws[f"B{row}"] = code
            ws[f"B{row}"].alignment = center

            # semicolon-separated indices
            ws[f"C{row}"] = "; ".join(sorted(indices, key=lambda x: int(x) if x.isdigit() else x))
            ws[f"C{row}"].alignment = left

            row += 1
    else:
        ws[f"A{row}"] = "No failed API endpoints."
        row += 1

    # ----- Auto-adjust column widths -----
    for col_idx in range(1, 6):  # up to column E for safety
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # Add some padding
        if max_length > 0:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)  # cap width at 60

    # ----- Save workbook -----
    wb.save(output_path)
    print(f"✅ Report generated at: {output_path}")
