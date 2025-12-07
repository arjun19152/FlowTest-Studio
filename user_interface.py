import sys
import json
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QFileDialog,
    QVBoxLayout, QPushButton, QWidget, QHBoxLayout, QMessageBox,
    QToolButton, QFrame, QSpacerItem, QSizePolicy, QMenu,
    QTabWidget, QDialog, QLineEdit, QDialogButtonBox, QScrollArea, QGridLayout
)
from PyQt5.QtGui import QIcon, QPixmap, QMouseEvent, QDrag
from PyQt5.QtCore import Qt, QPoint, QEvent, pyqtSignal, QMimeData
from PyQt5.QtGui import QCursor
from PyQt5 import QtGui, QtCore
import os
import pandas as pd
from utils import show_message
from openpyxl import load_workbook
from interactions import MainWindow
from main_backend import startEngine
from urllib.parse import urlparse
import shutil

# Custom Dialog for Scenario Name Input
class ScenarioNameDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Scenario")
        self.setFixedSize(300, 150)
        self.setStyleSheet("""
            QDialog {
                background-color: #3d4653;
                color: white;
                border: 1px solid #749BC2;
                border-radius: 5px;
            }
            QLabel {
                color: white;
                font-size: 14px;
            }
            QLineEdit {
                background-color: #2f3542;
                border: 1px solid #70a1ff;
                border-radius: 5px;
                padding: 5px;
                color: white;
                font-size: 14px;
                min-height: 30px;
            }
            QPushButton {
                background-color: #70a1ff;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1e90ff;
            }
            QPushButton:pressed {
                background-color: #007bff;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        self.name_label = QLabel("Scenario Name:")
        layout.addWidget(self.name_label)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("e.g., User Login Test")
        layout.addWidget(self.name_input)

        self.button_box = QDialogButtonBox()
        self.create_button = QPushButton("Create")
        self.button_box.addButton(self.create_button, QDialogButtonBox.AcceptRole)
        self.cancel_button = QPushButton("Cancel")
        self.button_box.addButton(self.cancel_button, QDialogButtonBox.RejectRole)

        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_scenario_name(self):
        return self.name_input.text().strip()

# New: Widget for displaying a single API name with an Add button
class ApiListItem(QFrame):
    add_api_requested = pyqtSignal(str)

    def __init__(self, api_name, parent=None):
        super().__init__(parent)
        self.api_name = api_name
        self.setFrameShape(QFrame.StyledPanel)
        self.setFrameShadow(QFrame.Raised)

        self.setStyleSheet("""
            ApiListItem {
                background-color: #3d4653;
                border: 1px solid #4a5364;
                border-radius: 5px;
                margin: 3px 0px; /* Adjusted margin to ensure it fits within ApiListPane padding */
            }
            ApiListItem QLabel {
                color: #e0e0e0;
                font-size: 13px;
                padding: 0px;
                background-color: transparent;
            }
            QPushButton.addApiButton {
                background-color: #5cb85c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 14px;
                font-weight: bold;
                min-width: 28px;
                min-height: 28px;
            }
            QPushButton.addApiButton:hover {
                background-color: #4cae4c;
            }
            QPushButton.addApiButton:pressed {
                background-color: #3f903f;
            }
        """)

        h_layout = QHBoxLayout(self)
        h_layout.setContentsMargins(8, 8, 8, 8)
        h_layout.setSpacing(10)

        self.api_label = QLabel(api_name)
        self.api_label.setWordWrap(True)
        h_layout.addWidget(self.api_label)
        h_layout.addStretch()

        self.add_button = QPushButton("+")
        self.add_button.setObjectName("addApiButton")
        self.add_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.add_button.clicked.connect(self._on_add_clicked)
        h_layout.addWidget(self.add_button)

    def _on_add_clicked(self):
        self.add_api_requested.emit(self.api_name)

# New: Left pane containing the list of APIs
class ApiListPane(QScrollArea):
    add_api_to_workspace = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.api_items = [] # To store references to ApiListItem widgets for filtering

        # Content widget that holds the actual layout and child widgets
        self.content_widget = QWidget()
        self.setWidget(self.content_widget)
        self.layout = QVBoxLayout(self.content_widget)
        # Apply padding here, so search bar and list items are inset
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.layout.setSpacing(5) # Spacing between API list items
        self.layout.setAlignment(Qt.AlignTop)

        # --- Add Search Bar ---
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Search APIs...")
        self.search_bar.setClearButtonEnabled(True)
        self.search_bar.textChanged.connect(self._filter_apis) # Connect signal for filtering
        # The search bar should be added directly to the main layout of the content_widget
        self.layout.addWidget(self.search_bar)
        # --- End Search Bar ---

        self.load_apis_from_config(os.path.join("configs","api_config_new.json"))
        self.apply_stylesheet() # Apply stylesheet after creating widgets

    def apply_stylesheet(self):
        self.setStyleSheet("""
            QScrollArea {
                background-color: #2f3542;
                border: none;
            }
            QScrollArea > QWidget { /* The content widget inside the scroll area */
                background-color: #2f3542;
            }
            QScrollBar:vertical {
                border: none;
                background: #57606f;
                width: 8px;
                margin: 0px 0px 0px 0px;
            }
            QScrollBar::handle:vertical {
                background: #749BC2;
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                background: none;
            }
            QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
                background: none;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }

            /* --- Search Bar Styling --- */
            ApiListPane QLineEdit { /* Target QLineEdit specifically within ApiListPane */
                background-color: #37404d; /* Slightly darker gray than pane background */
                border: 1px solid #4a5364; /* Similar border to ApiListItem */
                border-radius: 5px;
                padding: 6px 8px; /* Good padding for readability */
                color: #e0e0e0; /* Light gray text */
                font-size: 13px;
                margin-bottom: 5px; /* Space below search bar */
            }
            ApiListPane QLineEdit::placeholder {
                color: #a0a0a0; /* Muted placeholder color */
            }
            ApiListPane QLineEdit:focus {
                border: 1px solid #749BC2; /* Highlight border on focus */
                background-color: #3e4755; /* Slightly lighter background on focus */
            }
            /* --- End Search Bar Styling --- */
        """)

    def load_apis_from_config(self, config_file):
        # Clear existing API items, but keep the search bar
        for item in self.api_items:
            item.deleteLater() # Safely delete widgets
        self.api_items = []

        # Remove the stretch if it exists, to re-add after new items
        while self.layout.count() > 1 and isinstance(self.layout.itemAt(self.layout.count() - 1), QSpacerItem):
             # Ensure we don't remove the search bar (index 0)
            self.layout.takeAt(self.layout.count() - 1)

        try:
            with open(config_file, 'r') as file:
                api_data = json.load(file)

            for api_name in api_data.keys():
                api_item = ApiListItem(api_name)
                api_item.add_api_requested.connect(self.add_api_to_workspace.emit)
                self.layout.addWidget(api_item)
                self.api_items.append(api_item) # Store the item

            self.layout.addStretch() # Add stretch at the end to push items to top

        except FileNotFoundError:
            show_message("API Config Error", f"'{config_file}' not found. Please upload a project first.",level="warning")
        except json.JSONDecodeError:
            show_message("API Config Error", f"Error decoding JSON from '{config_file}'. Check file format.",level="critical")
        except Exception as e:
            show_message("API Config Error", f"An unexpected error occurred while loading API config: {str(e)}",level="critical")

    def _filter_apis(self, search_text):
        search_text_lower = search_text.lower()
        for api_item in self.api_items:
            if search_text_lower in api_item.api_name.lower():
                api_item.show()
            else:
                api_item.hide()

# New: Widget representing an API block in the main work area
class ApiBlock(QFrame):
    # Signals for removing
    remove_requested = pyqtSignal(QWidget)
    # Signals for reordering
    move_up_requested = pyqtSignal(QWidget)
    move_down_requested = pyqtSignal(QWidget)


    def __init__(self, api_name, method, url_path, parent=None):
        super().__init__(parent)
        self.api_name = api_name
        self.method = method
        self.url_path = url_path

        self.setFrameShape(QFrame.StyledPanel)
        self.setFrameShadow(QFrame.Raised)
        self.setFixedSize(300, 120)

        # --- MODIFIED CSS FOR ARROW BUTTONS & REVERTED REMOVE BUTTON ---
        self.normal_stylesheet = """
            ApiBlock {
                background-color: #57606f;
                border: 1px solid #749BC2;
                border-radius: 8px;
                margin: 5px;
            }
            ApiBlock:hover {
                border: 1px solid #a0c2e6; /* Lighter border on hover */
            }
            ApiBlock QLabel {
                color: white;
                font-size: 15px;
                font-weight: bold;
                padding: 0px;
                background-color: transparent;
            }
            ApiBlock QLabel#apiDetails {
                font-size: 13px;
                font-weight: normal;
                color: #e0e0e0;
                padding-top: 5px;
            }
            /* Original styling for remove button - REVERTED */
            ApiBlock QToolButton#removeButton {
                background-color: transparent;
                color: #ff6347; /* Set the 'x' color (e.g., a subtle red for delete) */
                border: none;
                border-radius: 0px;
                font-weight: bold;
                font-size: 16px;
                width: 20px;
                height: 20px;
                padding: 0px;
                margin: 0px;
            }
            ApiBlock QToolButton#removeButton:hover {
                color: #e65c3f; /* Slightly darker 'x' on hover */
                cursor: pointer;
            }

            /* New styling for arrow buttons */
            ApiBlock QToolButton#reorderButton {
                background-color: transparent;
                color: #ffffff; /* White color for arrows */
                border: none;
                border-radius: 4px; /* Slightly rounded corners */
                font-weight: bold;
                font-size: 16px; /* A bit bigger */
                width: 22px; /* A bit bigger */
                height: 22px; /* A bit bigger */
                padding: 0px;
                margin: 1px 0px; /* Increased vertical margin */
            }
            ApiBlock QToolButton#reorderButton:hover {
                background-color: #749BC2; /* Light blue background on hover */
                color: #ffffff; /* Keep text white on hover */
                cursor: pointer;
            }

            /* --- NEW: Styling for status labels --- */
            ApiBlock QLabel#statusPassed {
                background-color: #4CAF50; /* Green */
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 3px 8px;
                border-radius: 5px;
            }
            ApiBlock QLabel#statusFailed {
                background-color: #E43636; /* Red */
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 3px 8px;
                border-radius: 5px;
            }
        """
        self.highlight_stylesheet = """
            ApiBlock {
                background-color: #6a7485; /* Slightly lighter when dragging over */
                border: 2px dashed #a0c2e6; /* Dashed border for drop target */
                border-radius: 8px;
                margin: 5px;
            }
            ApiBlock QLabel { color: white; }
            ApiBlock QLabel#apiDetails { color: #e0e0e0; }
            ApiBlock QToolButton#removeButton {
                background-color: transparent;
                color: #ff6347;
                border: none;
            }
            ApiBlock QToolButton#reorderButton {
                background-color: transparent;
                color: #ffffff;
                border: none;
            }
        """
        self.setStyleSheet(self.normal_stylesheet)
        # --- END MODIFIED CSS ---

        # Main vertical layout for content
        v_layout = QVBoxLayout(self)
        v_layout.setContentsMargins(10, 10, 10, 10)
        v_layout.setSpacing(5)

        # Horizontal layout for API name, close button, and new arrow buttons
        top_h_layout = QHBoxLayout()
        top_h_layout.setContentsMargins(0, 0, 0, 0)
        top_h_layout.setSpacing(5)

        self.name_label = QLabel(f"API: {api_name}")
        self.name_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.name_label.setWordWrap(True)
        top_h_layout.addWidget(self.name_label)
        top_h_layout.addStretch() # Pushes everything to the right

        # --- NEW: Arrow buttons layout ---
        arrow_buttons_v_layout = QVBoxLayout()
        arrow_buttons_v_layout.setContentsMargins(0, 0, 0, 0)
        arrow_buttons_v_layout.setSpacing(2) # Increased spacing between arrow buttons

        # Up arrow button
        self.up_button = QToolButton(self)
        self.up_button.setText("▲") # Unicode up arrow
        self.up_button.setObjectName("reorderButton") # New object name for reorder buttons
        self.up_button.clicked.connect(self._on_up_clicked)
        self.up_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        arrow_buttons_v_layout.addWidget(self.up_button)

        # Down arrow button
        self.down_button = QToolButton(self)
        self.down_button.setText("▼") # Unicode down arrow
        self.down_button.setObjectName("reorderButton") # New object name for reorder buttons
        self.down_button.clicked.connect(self._on_down_clicked)
        self.down_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        arrow_buttons_v_layout.addWidget(self.down_button)

        # Add the arrow buttons layout to the top horizontal layout
        top_h_layout.addLayout(arrow_buttons_v_layout)
        # --- END NEW ---

        # Close button (keep it to the right of arrows)
        self.close_button = QToolButton(self)
        self.close_button.setText("x")
        self.close_button.setObjectName("removeButton")  # Set object name for specific styling
        self.close_button.clicked.connect(self._on_close_clicked)
        self.close_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        top_h_layout.addWidget(self.close_button)


        v_layout.addLayout(top_h_layout)  # Add the top horizontal layout first

        self.details_label = QLabel(f"Method: {self.method}, URL: {self.url_path}")
        self.details_label.setObjectName("apiDetails")
        self.details_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        v_layout.addWidget(self.details_label)

        v_layout.addStretch()

        # --- NEW: Status label layout at the bottom right ---
        self.status_label_pass = QLabel()
        self.status_label_pass.setAlignment(Qt.AlignBottom | Qt.AlignRight)
        self.status_label_pass.hide()  # Initially hide the label

        self.status_label_fail = QLabel()
        self.status_label_fail.setAlignment(Qt.AlignBottom | Qt.AlignRight)
        self.status_label_fail.hide()  # Initially hide the label

        bottom_h_layout = QHBoxLayout()
        bottom_h_layout.addStretch() # Pushes the label to the right
        bottom_h_layout.addWidget(self.status_label_pass)
        bottom_h_layout.addWidget(self.status_label_fail)
        
        v_layout.addLayout(bottom_h_layout)
        # --- END NEW ---

    def _on_close_clicked(self):
        self.remove_requested.emit(self)  # Emit signal with self as argument

    def _on_up_clicked(self):
        """Handler for the up arrow button click."""
        self.move_up_requested.emit(self)

    def _on_down_clicked(self):
        """Handler for the down arrow button click."""
        self.move_down_requested.emit(self)

    # MODIFIED: mousePressEvent to only handle button clicks
    def mousePressEvent(self, event: QMouseEvent):
        # Allow buttons to handle their own clicks, prevent propagation to parent
        if self.close_button.geometry().contains(event.pos()) or \
           self.up_button.geometry().contains(event.pos()) or \
           self.down_button.geometry().contains(event.pos()):
            event.ignore()
        else:
            # Optionally, you could still call super() if you want other default QFrame mouse press behavior
            # For example, if you want the frame to receive focus on click, or similar.
            super().mousePressEvent(event)

    def set_progress(self, text: str):
        self.status_label_pass.setObjectName("")
        self.status_label_pass.setText(text)
        self.status_label_pass.show()

        # Hide failed label only during progress
        self.status_label_fail.setVisible(False)

        QApplication.processEvents()


    def set_status(self, success, fail):

        self.status_label_pass.setText("Passed: "+str(success))
        self.status_label_pass.setObjectName("statusPassed")

        self.status_label_pass.show()
        self.status_label_pass.style().polish(self.status_label_pass)
        self.status_label_pass.repaint()

        if fail!=0:
            self.status_label_fail.setText("Failed: "+str(fail))
            self.status_label_fail.setObjectName("statusFailed")

            self.status_label_fail.show()
            self.status_label_fail.style().polish(self.status_label_fail)
            self.status_label_fail.repaint()

        QApplication.processEvents()

    def reset_status(self):
        """
        Resets the API block status label before a new execution starts.
        """
        self.status_label_pass.setText("")  # Clear text
        self.status_label_pass.setObjectName("")  # Remove style classes
        self.status_label_pass.hide()  # Hide until used
        self.status_label_pass.style().polish(self.status_label_pass)  # Refresh style
        self.status_label_pass.repaint()

        self.status_label_fail.setText("")  # Clear text
        self.status_label_fail.setObjectName("")  # Remove style classes
        self.status_label_fail.hide()  # Hide until used
        self.status_label_fail.style().polish(self.status_label_fail)  # Refresh style
        self.status_label_fail.repaint()



# Main application window class
class FlowTestStudio(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setGeometry(100, 100, 900, 600)
        self.setStyleSheet(self.load_stylesheet())
        self.drag_position = None # This is for window drag, keep it.

        self.initial_client_area = None
        self.workspace_widget = None
        self.client_menu_bar = None
        self.scenario_tabs = None
        self.search_bar = None
        self.api_list_pane = None
        self.scenario_name = None

        self.current_tab_work_layout = None
        self.current_tab_work_area_content_widget = None

        self.current_project_file = None

        # Keep track of occupied cells in the grid for each tab
        self.tab_grid_occupancy = {}
        self.on_change = False
        self.init_ui()


    def load_stylesheet(self):
        return """
        QMainWindow {
            background-color: #2f3542;
            border: 1px solid #749BC2;
            border-radius: 5px;
        }
        QFrame#ClientMenuBar {
                background-color: #3d4653;
                border-bottom: 1px solid #4a5364;
        }
        QLabel {
            color: white;
        }
        QPushButton {
            background-color: #70a1ff;
            color: white;
            padding: 10px 20px;
            border-radius: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #1e90ff;
        }
        QToolButton {
            background-color: transparent;
            padding: 6px 12px;
            border: none;
            border-radius: 6px;
            color: white;
        }
        QToolButton:hover {
            background-color: #70a1ff;
            border: 1px solid white;
        }
        QMenu {
            background-color: #57606f;
            color: white;
            padding: 4px;
        }
        QMenu::item:selected {
            background-color: #70a1ff;
        }
        QFrame#ClientMenuBar {
            background-color: #D7D7D7;
            border-bottom: none;
            border-top: none;
            border-left: none;
            border-right: none;
        }

        QTabWidget::pane {
            border: 1px solid #57606f;
            background-color: #2f3542;
        }

        QTabBar::tab {
            background: #57606f;
            color: white;
            padding: 8px 20px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            margin-right: 2px;
        }

        QTabBar::tab:selected {
            background: #749BC2;
            color: white;
            border-bottom: 2px solid #749BC2;
        }

        QTabBar::tab:hover {
            background: #70a1ff;
        }
        """

    def init_ui(self):
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        self.main_layout = QVBoxLayout(self.main_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        self.menu_bar = self.create_custom_menu_bar()
        self.main_layout.addWidget(self.menu_bar)

        self.show_initial_client_area()

    def _on_scenario_search_changed(self, text):
        """Finds the first scenario tab matching the search text and brings it into focus."""
        # Ensure workspace and tabs exist before searching
        if not self.workspace_widget or not self.scenario_tabs:
            return
            
        # Ignore search if the text is empty or is the placeholder
        if not text or text == self.search_bar._placeholder_text:
            return

        search_text_lower = text.lower()

        # Find the first tab whose name contains the search text (case-insensitive)
        for i in range(self.scenario_tabs.count()):
            tab_title = self.scenario_tabs.tabText(i)
            if search_text_lower in tab_title.lower():
                self.scenario_tabs.setCurrentIndex(i)
                return # Stop after focusing the first match

    
    def show_initial_client_area(self):
        if self.workspace_widget and self.workspace_widget.isVisible():
            self.workspace_widget.hide()
            self.main_layout.removeWidget(self.workspace_widget)

        if self.client_menu_bar and self.client_menu_bar.isVisible():
            self.client_menu_bar.hide()
            self.main_layout.removeWidget(self.client_menu_bar)

        if not self.initial_client_area:
            self.initial_client_area = QWidget()
            self.initial_client_area.setStyleSheet("background-color: #2f3542; color: white;")

            layout = QVBoxLayout()
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setAlignment(Qt.AlignCenter)

            logo_label = QLabel()
            pixmap = QPixmap("logo.png")
            if pixmap.isNull():
                print("Error: logo.png not found or could not be loaded. Please ensure 'logo.png' is in the same directory.")
                logo_label.setText("Logo")
                logo_label.setStyleSheet("color: white; font-size: 24px;")
            else:
                logo_label.setPixmap(pixmap.scaled(120, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            logo_label.setAlignment(Qt.AlignCenter)

            create_btn = QPushButton("Create New Project")
            create_btn.setFixedSize(220, 50)
            create_btn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
            create_btn.setStyleSheet("""
                QPushButton {
                    background-color: #578FCA;
                    color: white;
                    padding: 10px 20px;
                    border: 2px solid #578FCA;
                    border-radius: 10px;
                    font-size: 16px;
                    box-shadow: 2px 2px 5px rgba(0, 0, 0, 0.3);
                    cursor: arrow;
                }
                QPushButton:hover {
                    background-color: #3674B5;
                    border: 2px solid #3674B5;
                    box-shadow: 3px 3px 8px rgba(0, 0, 0, 0.5);
                    cursor: pointer ;
                }
                QPushButton:pressed {
                    background-color: #3674B5;
                    border: 2px solid #3674B5;
                    box-shadow: inset 1px 1px 3px rgba(0, 0, 0, 0.4);
                }
            """)

            create_btn.clicked.connect(self.create_new_project)

            layout.addWidget(logo_label)
            layout.addSpacing(10)
            layout.addWidget(create_btn)

            self.initial_client_area.setLayout(layout)
            self.main_layout.addWidget(self.initial_client_area)
        else:
            self.main_layout.addWidget(self.initial_client_area)
            self.initial_client_area.show()

    def show_workspace(self):
        if self.initial_client_area and self.initial_client_area.isVisible():
            self.initial_client_area.hide()
            self.main_layout.removeWidget(self.initial_client_area)

        if not self.client_menu_bar:
            self.client_menu_bar = self.create_client_menu_bar()
            self.main_layout.insertWidget(1, self.client_menu_bar)
        self.client_menu_bar.show()

        if not self.workspace_widget:
            self.workspace_widget = QWidget()
            workspace_layout = QVBoxLayout(self.workspace_widget)
            workspace_layout.setContentsMargins(0, 0, 0, 0)
            workspace_layout.setSpacing(0)

            self.scenario_tabs = QTabWidget()
            self.scenario_tabs.setTabsClosable(True)
            self.scenario_tabs.tabCloseRequested.connect(self.close_scenario_tab)
            self.scenario_tabs.setMovable(True)
            self.scenario_tabs.currentChanged.connect(self._update_current_tab_work_layout)

            workspace_layout.addWidget(self.scenario_tabs)

            self.main_layout.addWidget(self.workspace_widget)

        self.workspace_widget.show()
        self.setFocus()

    def create_custom_menu_bar(self):
        menu_bar = QFrame()
        menu_bar.setFixedHeight(40)
        menu_bar.setStyleSheet("background-color: #57606f;")

        layout = QHBoxLayout()
        layout.setContentsMargins(8, 0, 8, 0)
        layout.setSpacing(10)

        logo_label = QLabel()
        logo_pixmap = QPixmap("logo.png")
        if logo_pixmap.isNull():
            print("Error: logo.png not found or could not be loaded. Please ensure 'logo.png' is in the same directory.")
            logo_label.setText("Logo")
            logo_label.setStyleSheet("color: white; font-size: 14px;")
        else:
            logo_label.setPixmap(logo_pixmap.scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        layout.addWidget(logo_label)

        file_btn = QToolButton()
        file_btn.setText("File")
        file_btn.setPopupMode(QToolButton.InstantPopup)
        file_btn.setStyleSheet("QToolButton::menu-indicator { image: none; }")

        file_menu = QMenu()
        file_menu.addAction("New Project", self.create_new_project)
        file_menu.addAction("Open Project", self.open_project)
        file_menu.addAction("Save", self.save_project)
        file_menu.addAction("Save As", self.save_project_as)

        file_btn.setMenu(file_menu)
        layout.addWidget(file_btn)

        layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))

        minimize = QToolButton()
        minimize.setText("–")
        minimize.setStyleSheet("font-size: 17px;")
        minimize.clicked.connect(self.showMinimized)
        layout.addWidget(minimize)

        maximize = QToolButton()
        maximize.setText("▭")
        maximize.setStyleSheet("font-size: 17px;")
        maximize.clicked.connect(lambda: self.showNormal() if self.isMaximized() else self.showMaximized())
        layout.addWidget(maximize)

        close = QToolButton()
        close.setText("x")
        close.setStyleSheet("font-size: 16px; padding-top: 5px;")
        close.clicked.connect(self.close)
        layout.addWidget(close)

        menu_bar.setLayout(layout)
        return menu_bar

    def upload_testcases(self):
        """Handle test case file upload"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Upload Test Cases", 
            "", 
            "Excel Files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
            
        try:
            # Parse the Excel file
            testcases = {}
            wb = load_workbook(file_path)
            sheet_names = wb.sheetnames
            
            for sheet_name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                scenario_data = {}
                
                for col in df.columns:
                    api_name = str(df[col].iloc[0]).strip()
                    if not api_name:
                        continue
                        
                    test_cases = []
                    for row in range(1, len(df)):
                        test_case_str = str(df[col].iloc[row]).strip()
                        if test_case_str and test_case_str != 'nan':
                            try:
                                test_case = json.loads(test_case_str.replace("'", '"'))
                                test_cases.append(test_case)
                            except json.JSONDecodeError:
                                show_message("Invalid JSON", f"Invalid JSON in sheet {sheet_name}, API {api_name}, row {row+1}",level="warning")
                                continue
                                
                    if test_cases:
                        scenario_data[api_name] = test_cases
                
                if scenario_data:
                    testcases[sheet_name] = scenario_data
                    
            # Save to JSON file
            os.makedirs("testcases", exist_ok=True)
            output_path = os.path.join("testcases", "testcases.json")
            with open(output_path, 'w') as f:
                json.dump(testcases, f, indent=4)
                
            show_message("Success", f"Successfully uploaded and saved {len(testcases)} scenario(s) to {output_path}",level="info")
            
        except PermissionError:
            show_message("Error", "Please close the Excel file before uploading",level="critical")
        except Exception as e:
            show_message("Error", f"Failed to process Excel file: {str(e)}",level="critical")


    def create_client_menu_bar(self):
        client_menu_bar = QFrame()
        client_menu_bar.setFixedHeight(35)
        client_menu_bar.setObjectName("ClientMenuBar")

        layout = QHBoxLayout()
        layout.setContentsMargins(8, 0, 8, 0)
        layout.setSpacing(10)

        self.search_bar = QLineEdit() # This is the main scenario search bar
        self.search_bar.setClearButtonEnabled(True)
        self.search_bar.setFixedWidth(350)
        
        # --- NEW: Connect the textChanged signal ---
        self.search_bar.textChanged.connect(self._on_scenario_search_changed)
        # --- End New ---

        self.search_bar._placeholder_text = "Search scenarios..."
        self.search_bar._placeholder_color = "#819A91"
        self.search_bar._active_text_color = "#2d3f85"

        self.search_bar.setText(self.search_bar._placeholder_text)
        self.search_bar.setStyleSheet(f"""
            QLineEdit {{
                background-color: #D7D7D7;
                border: 1px solid #B1AB86;
                border-radius: 8px;
                padding: 5px 15px;
                padding-left: 15px;
                color: {self.search_bar._placeholder_color};
                selection-background-color: #70a1ff;
                selection-color: white;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border: 1px solid #819A91;
                background-color: #D7D7D7;
            }}
        """)

        search_icon_pixmap = QPixmap("search_icon.png")
        if not search_icon_pixmap.isNull():
            scaled_search_icon = search_icon_pixmap.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.search_bar.addAction(QIcon(scaled_search_icon), QLineEdit.LeadingPosition)

        layout.addWidget(self.search_bar)

        self.search_bar.installEventFilter(self)
        self.search_bar.clearFocus()

        layout.addStretch()

        # ... (the rest of the method remains unchanged)
        # ... (add_scenario_btn, run_scenario_btn, etc.)
        
        add_scenario_btn = QPushButton("+ Add Scenario")
        add_scenario_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 6px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        add_scenario_btn.clicked.connect(self.add_scenario_work_section)
        layout.addWidget(add_scenario_btn)

        add_interaction_btn = QPushButton("+ Add Interaction")
        add_interaction_btn.setToolTip("Add interactions to active scenario")
        add_interaction_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 6px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        add_interaction_btn.clicked.connect(self.launch_interaction_window)
        layout.addWidget(add_interaction_btn)

        run_scenario_btn = QPushButton(">> Run")
        run_scenario_btn.setToolTip("Run active scenario")
        run_scenario_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 6px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        run_scenario_btn.clicked.connect(self.on_run_backend)
        layout.addWidget(run_scenario_btn)

        run_all_scenario_btn = QPushButton(">> Run all")
        run_all_scenario_btn.setToolTip("Run all scenario")
        run_all_scenario_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 6px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        layout.addWidget(run_all_scenario_btn)

        upload_testcases_btn = QPushButton("Testcases")
        upload_testcases_btn.setToolTip("Upload all testcases")
        original_pixmap = QPixmap("upload.png")
        scaled_pixmap = original_pixmap.scaled(14, 14, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        upload_testcases_btn.setIcon(QIcon(scaled_pixmap))
        upload_testcases_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 4px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        upload_testcases_btn.clicked.connect(self.upload_testcases)
        layout.addWidget(upload_testcases_btn)

        env_variables_btn = QPushButton("+ Env variables")
        env_variables_btn.setToolTip("Add environment variables")
        env_variables_btn.setStyleSheet("""
            QPushButton {
                background-color: #328E6E; color: white; border: none;
                border-radius: 5px; padding: 6px 12px; font-size: 13px;
            }
            QPushButton:hover { background-color: #3F7D58; }
        """)
        layout.addWidget(env_variables_btn)

        client_menu_bar.setLayout(layout)
        return client_menu_bar

    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton:
            current_widget = self.childAt(event.pos())
            if (current_widget == self.menu_bar or \
               (self.initial_client_area and current_widget == self.initial_client_area) or \
               (self.workspace_widget and current_widget == self.workspace_widget) or \
               (self.client_menu_bar and current_widget == self.client_menu_bar)) and \
               current_widget != self.search_bar: # Make sure the drag doesn't interfere with main search bar
                self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
                event.accept()
            else:
                super().mousePressEvent(event)
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent):
        if event.buttons() == Qt.LeftButton and self.drag_position:
            self.move(event.globalPos() - self.drag_position)
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):
        self.drag_position = None
        super().mouseReleaseEvent(event)

    def eventFilter(self, obj, event):
        if obj == self.search_bar: # This handles the *main* search bar's placeholder text
            if event.type() == QEvent.FocusIn:
                if self.search_bar.text() == self.search_bar._placeholder_text:
                    self.search_bar.clear()
                    current_style = self.search_bar.styleSheet()
                    new_style = current_style.replace(
                        f"color: {self.search_bar._placeholder_color};",
                        f"color: {self.search_bar._active_text_color};"
                    )
                    self.search_bar.setStyleSheet(new_style)

            elif event.type() == QEvent.FocusOut:
                if self.search_bar.text().strip() == "":
                    self.search_bar.setText(self.search_bar._placeholder_text)
                    current_style = self.search_bar.styleSheet()
                    new_style = current_style.replace(
                        f"color: {self.search_bar._active_text_color};",
                        f"color: {self.search_bar._placeholder_color};"
                    )
                    self.search_bar.setStyleSheet(new_style)
        return super().eventFilter(obj, event)

    def create_new_project(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Upload API JSON File", "", "JSON Files (*.json)")
        if file_path:
            try:
                with open(file_path, 'r') as file:
                    data = json.load(file)
                    transformed = self.transform_json(data)
                    with open(os.path.join("configs","api_config_new.json"), "w") as outfile:
                        json.dump(transformed, outfile, indent=4)

                    self.show_workspace()

                    # New: Clear existing tabs and ensure a fresh start
                    for i in reversed(range(self.scenario_tabs.count())):
                        self.close_scenario_tab(i)

                    # Reset current project file path, as this is a *new* project
                    self.current_project_file = None
                    show_message("New Project", "New project initiated. Please add a scenario to begin.",level="info")


                    # Ensure api_list_pane is reloaded after a new project is created
                    if self.scenario_tabs.currentWidget(): # Check if a tab exists
                        current_tab_layout = self.scenario_tabs.currentWidget().layout()
                        if current_tab_layout and isinstance(current_tab_layout.itemAt(0).widget(), ApiListPane):
                            self.api_list_pane = current_tab_layout.itemAt(0).widget()
                            self.api_list_pane.load_apis_from_config(os.path.join("configs","api_config_new.json"))
                        # else:
                            # This case implies no active tab or a different layout.
                            # For simplicity, we'll just show the workspace,
                            # and the api_list_pane will be created/loaded when
                            # a new scenario is added.

            except FileNotFoundError:
                show_message("Error", f"File not found: {file_path}",level="critical")
                print("This")
            except json.JSONDecodeError:
                show_message("Error", f"Invalid JSON file: {file_path}. Please check the file format.",level="critical")
                print("his may")
            except Exception as e:
                show_message("Error", f"An unexpected error occurred: {str(e)}",level="critical")
                print("or this")

    def open_project(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Project", "", "JSON Files (*.json)")
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    project_data = json.load(f)

                print("Project loaded.")                    
                # Clear existing tabs before loading new ones
                if self.scenario_tabs and self.scenario_tabs.count() > 0:
                    for i in reversed(range(self.scenario_tabs.count())):
                        self.close_scenario_tab(i)

                print("All previously loaded tabs closed")
                self.show_workspace()
                
                if "scenarios" in project_data and isinstance(project_data["scenarios"], dict):
                    for scenario_name, api_names_list in project_data["scenarios"].items():

                        print(scenario_name)
                        print(api_names_list)
                        # Create a new tab for each scenario
                        self.add_scenario_work_section_new(scenario_name=scenario_name)
                        for api_name in api_names_list:
                            print("API name: ", api_name)
                            self.add_api_block_to_current_tab(api_name) 
                

                self.current_project_file = file_path # Set the current project file
                show_message("Opened", f"Project loaded from: {file_path}",level="info")

                # Similar to create_new_project, ensure the API list is reloaded
                if self.scenario_tabs.currentWidget():
                    current_tab_layout = self.scenario_tabs.currentWidget().layout()
                    if current_tab_layout and isinstance(current_tab_layout.itemAt(0).widget(), ApiListPane):
                        self.api_list_pane = current_tab_layout.itemAt(0).widget()
                        self.api_list_pane.load_apis_from_config(os.path.join("configs","api_config_new.json"))

            except FileNotFoundError:
                show_message("Error", f"Project file not found: {file_path}",level="critical")
            except json.JSONDecodeError:
                show_message("Error", f"Invalid JSON project file: {file_path}. Please check the file format.",level="critical")
            except Exception as e:
                show_message("Error", f"An error occurred while opening the project: {str(e)}",level="critical")

            
    def add_scenario_work_section(self):
        dialog = ScenarioNameDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            scenario_name = dialog.get_scenario_name()
            
            if scenario_name:
                self.scenario_name = scenario_name
                for i in range(self.scenario_tabs.count()):
                    if self.scenario_tabs.tabText(i) == scenario_name:
                        show_message("Duplicate Scenario",f"A scenario named '{scenario_name}' already exists.",level="warning")
                        self.scenario_tabs.setCurrentIndex(i)
                        return

                new_tab_page = QWidget()
                page_main_layout = QHBoxLayout(new_tab_page)
                page_main_layout.setContentsMargins(0, 0, 0, 0)
                page_main_layout.setSpacing(0)
                new_tab_page.setStyleSheet("background-color: #2f3542;")

                self.api_list_pane = ApiListPane()
                self.api_list_pane.setFixedWidth(250)
                self.api_list_pane.add_api_to_workspace.connect(self.add_api_block_to_current_tab)
                page_main_layout.addWidget(self.api_list_pane)

                self.current_tab_work_area_container = QFrame()
                self.current_tab_work_area_container.setStyleSheet("background-color: #3d4653; border-left: 1px solid #57606f;")
                self.current_tab_work_area_container.setFrameShape(QFrame.NoFrame)
                self.current_tab_work_area_container.setFrameShadow(QFrame.Plain)

                container_layout = QVBoxLayout(self.current_tab_work_area_container)
                container_layout.setContentsMargins(0, 0, 0, 0)
                container_layout.setSpacing(0)

                work_scroll_area = QScrollArea()
                work_scroll_area.setWidgetResizable(True)
                work_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
                work_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                work_scroll_area.setStyleSheet("""
                    QScrollArea {
                        border: none;
                        background-color: #3d4653;
                    }
                    QScrollArea > QWidget {
                        background-color: #3d4653;
                    }
                    QScrollBar:vertical {
                        border: none;
                        background: #57606f;
                        width: 8px;
                        margin: 0px 0px 0px 0px;
                    }
                    QScrollBar::handle:vertical {
                        background: #749BC2;
                        border-radius: 4px;
                        min-height: 20px;
                    }
                    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                        background: none;
                    }
                    QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
                        background: none;
                    }
                    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                        background: none;
                    }
                """)

                self.current_tab_work_area_content_widget = QWidget()
                self.current_tab_work_area_content_widget.setStyleSheet("background-color: #3d4653;")

                # REPLACED QVBoxLayout with QGridLayout (8x4)
                self.current_tab_work_layout = QGridLayout(self.current_tab_work_area_content_widget)
                self.current_tab_work_layout.setContentsMargins(10, 10, 10, 10)
                self.current_tab_work_layout.setSpacing(10)
                # Align to top-left is default for QGridLayout; we don't need addStretch() for grids
                # to fill empty space, but you can set stretch factors for rows/columns if needed.

                # Initialize grid occupancy for this new tab
                self.tab_grid_occupancy[self.scenario_tabs.count()] = set() # Use tab index as key

                work_scroll_area.setWidget(self.current_tab_work_area_content_widget)
                container_layout.addWidget(work_scroll_area)
                page_main_layout.addWidget(self.current_tab_work_area_container, 1)

                index = self.scenario_tabs.addTab(new_tab_page, scenario_name)
                self.scenario_tabs.setCurrentIndex(index)

                self._update_current_tab_work_layout()
                self.api_list_pane.load_apis_from_config(os.path.join("configs","api_config_new.json"))

            else:
                show_message("Invalid Name","Scenario name cannot be empty.",level="warning")
                

    def add_scenario_work_section_new(self, scenario_name):
        if scenario_name:
            
            new_tab_page = QWidget()
            page_main_layout = QHBoxLayout(new_tab_page)
            page_main_layout.setContentsMargins(0, 0, 0, 0)
            page_main_layout.setSpacing(0)
            new_tab_page.setStyleSheet("background-color: #2f3542;")

            self.api_list_pane = ApiListPane()
            self.api_list_pane.setFixedWidth(250)
            self.api_list_pane.add_api_to_workspace.connect(self.add_api_block_to_current_tab)
            page_main_layout.addWidget(self.api_list_pane)

            self.current_tab_work_area_container = QFrame()
            self.current_tab_work_area_container.setStyleSheet("background-color: #3d4653; border-left: 1px solid #57606f;")
            self.current_tab_work_area_container.setFrameShape(QFrame.NoFrame)
            self.current_tab_work_area_container.setFrameShadow(QFrame.Plain)

            container_layout = QVBoxLayout(self.current_tab_work_area_container)
            container_layout.setContentsMargins(0, 0, 0, 0)
            container_layout.setSpacing(0)

            work_scroll_area = QScrollArea()
            work_scroll_area.setWidgetResizable(True)
            work_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            work_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            work_scroll_area.setStyleSheet("""
                QScrollArea {
                    border: none;
                    background-color: #3d4653;
                }
                QScrollArea > QWidget {
                    background-color: #3d4653;
                }
                QScrollBar:vertical {
                    border: none;
                    background: #57606f;
                    width: 8px;
                    margin: 0px 0px 0px 0px;
                }
                QScrollBar::handle:vertical {
                    background: #749BC2;
                    border-radius: 4px;
                    min-height: 20px;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    background: none;
                }
                QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
                    background: none;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }
            """)

            self.current_tab_work_area_content_widget = QWidget()
            self.current_tab_work_area_content_widget.setStyleSheet("background-color: #3d4653;")

            # REPLACED QVBoxLayout with QGridLayout (8x4)
            self.current_tab_work_layout = QGridLayout(self.current_tab_work_area_content_widget)
            self.current_tab_work_layout.setContentsMargins(10, 10, 10, 10)
            self.current_tab_work_layout.setSpacing(10)
            # Align to top-left is default for QGridLayout; we don't need addStretch() for grids
            # to fill empty space, but you can set stretch factors for rows/columns if needed.

            # Initialize grid occupancy for this new tab
            self.tab_grid_occupancy[self.scenario_tabs.count()] = set()


            work_scroll_area.setWidget(self.current_tab_work_area_content_widget)
            container_layout.addWidget(work_scroll_area)
            page_main_layout.addWidget(self.current_tab_work_area_container, 1)

            index = self.scenario_tabs.addTab(new_tab_page, scenario_name)
            self.scenario_tabs.setCurrentIndex(index)

            self._update_current_tab_work_layout()
            self.api_list_pane.load_apis_from_config(os.path.join("configs","api_config_new.json"))

            self.scenario_name = scenario_name

        else:
            show_message("Invalid Name","Scenario name cannot be empty.",level="warning")


    def _update_current_tab_work_layout(self):
        current_widget = self.scenario_tabs.currentWidget()
        if current_widget:
            page_layout = current_widget.layout()
            if page_layout and page_layout.count() > 1:
                api_pane_widget = page_layout.itemAt(0).widget()
                if isinstance(api_pane_widget, ApiListPane):
                    self.api_list_pane = api_pane_widget
                else:
                    self.api_list_pane = None

                work_area_container = page_layout.itemAt(1).widget()
                if isinstance(work_area_container, QFrame):
                    container_layout = work_area_container.layout()
                    if container_layout and container_layout.count() > 0:
                        work_scroll_area = container_layout.itemAt(0).widget()
                        if isinstance(work_scroll_area, QScrollArea):
                            self.current_tab_work_area_content_widget = work_scroll_area.widget()
                            # IMPORTANT: Ensure this is now a QGridLayout
                            if isinstance(self.current_tab_work_area_content_widget.layout(), QGridLayout):
                                self.current_tab_work_layout = self.current_tab_work_area_content_widget.layout()
                            else:
                                # This case should ideally not happen if add_scenario_work_section is consistent
                                self.current_tab_work_layout = None
                                print("Warning: Expected QGridLayout but found a different layout.")
                        else:
                            self.current_tab_work_layout = None
                            self.current_tab_work_area_content_widget = None
                    else:
                        self.current_tab_work_layout = None
                        self.current_tab_work_area_content_widget = None
                else:
                    self.current_tab_work_layout = None
                    self.current_tab_work_area_content_widget = None

            else:
                self.current_tab_work_layout = None
                self.current_tab_work_area_content_widget = None
                self.api_list_pane = None
        else:
            self.current_tab_work_layout = None
            self.current_tab_work_area_content_widget = None
            self.api_list_pane = None

    def add_api_block_to_current_tab(self, api_name):
        self.on_change = True
        if self.current_tab_work_layout.count() < 0:
            show_message("No Active Scenario","Please create or select a scenario tab first.",level="warning")
            return

        api_data = {}
        try:
            # Using a fixed filename for api_config, as per original logic
            with open(os.path.join("configs","api_config_new.json"), 'r') as file:
                full_config = json.load(file)
                api_data = full_config.get(api_name, {})
        except (FileNotFoundError, json.JSONDecodeError):
            show_message("Error","API configuration not found or invalid.",level="warning")
            return

        method = api_data.get("method", "UNKNOWN")
        url_path = api_data.get("path", "/example/path")

        api_block = ApiBlock(api_name, method, url_path)
        api_block.remove_requested.connect(self._remove_api_block)
        api_block.move_up_requested.connect(self._move_api_block_up)
        api_block.move_down_requested.connect(self._move_api_block_down)

        current_tab_index = self.scenario_tabs.currentIndex()
        if current_tab_index not in self.tab_grid_occupancy:
            self.tab_grid_occupancy[current_tab_index] = set()

        # Find the next available cell (8x4 grid)
        added = False
        for col in range(4):
            for row in range(8):
                if (row, col) not in self.tab_grid_occupancy[current_tab_index]:
                    self.current_tab_work_layout.addWidget(api_block, row, col)
                    self.tab_grid_occupancy[current_tab_index].add((row, col))
                    api_block.setProperty("grid_pos", (row, col)) # Store position for easy lookup
                    added = True
                    print(f"Added API block for: {api_name} at ({row}, {col})")
                    break
            if added:
                break

        if not added:
            show_message("Grid Full","The 8x4 grid is full. Please remove existing API blocks.",level="warning")


    def _remove_api_block(self, api_block_widget: QWidget):
        self.on_change = True
        if self.current_tab_work_layout:
            # Remove from grid occupancy
            current_tab_index = self.scenario_tabs.currentIndex()
            if current_tab_index in self.tab_grid_occupancy:
                grid_pos = api_block_widget.property("grid_pos")
                if grid_pos and grid_pos in self.tab_grid_occupancy[current_tab_index]:
                    self.tab_grid_occupancy[current_tab_index].remove(grid_pos)

            self.current_tab_work_layout.removeWidget(api_block_widget)
            api_block_widget.deleteLater()
            print(f"Removed API block: {api_block_widget.api_name}")


    def close_scenario_tab(self, index):
        if index >= 0:
            # Remove grid occupancy data for the closed tab
            if index in self.tab_grid_occupancy:
                del self.tab_grid_occupancy[index]

            widget_to_remove = self.scenario_tabs.widget(index)
            self.scenario_tabs.removeTab(index)
            widget_to_remove.deleteLater()

            if self.scenario_tabs.count() == 0:
                print("No scenarios left.")
                self.api_list_pane = None
                self.current_tab_work_layout = None
                self.current_tab_work_area_content_widget = None
                # self.dragged_api_block = None # This variable doesn't seem to be part of FlowTestStudio anymore.
            else:
                self._update_current_tab_work_layout()



    def transform_json(self, data):
        output = {}
        for item in data.get("item", []):
            name = item.get("name", "Unnamed_API")
            request = item.get("request", {})
            url_data = request.get("url", {})
            method = request.get("method", "GET").upper()
            headers = {h["key"]: h["value"] for h in request.get("header", []) if "key" in h and "value" in h}
            body = {}

            if request.get("body", {}).get("mode") == "raw":
                try:
                    body = json.loads(request["body"]["raw"])
                except json.JSONDecodeError:
                    body = {}

            full_url = ""
            path = "/"
            params = {}

            if isinstance(url_data, dict):
                # Use raw URL for parsing
                full_url = url_data.get("raw", "")
                # Extract query parameters from Postman's format
                query_list = url_data.get("query", [])
                for q in query_list:
                    params[q["key"]] = q["value"]
            else:
                full_url = url_data

            if full_url:
                parsed_url = urlparse(full_url)
                # The base URL is the scheme + netloc (e.g., https://alpytoavvqadpdasdlez.supabase.co)
                base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
                # The path is just the path part of the URL (e.g., /rest/v1/users)
                path = parsed_url.path
                
                # The logic to remove /rest/v1 is now removed, so the full path is kept.
                if not path:
                    path = '/'
            else:
                base_url = ""

            output[name] = {
                "url": base_url,
                "method": method,
                "headers": headers,
                "path": path  # The full path is now stored here
            }

            if params:
                output[name]["params"] = params
            if body:
                output[name]["body"] = body

        return output



    def _move_api_block_up(self, api_block_widget: QWidget):
        self.on_change = True
        if not self.current_tab_work_layout:
            return

        current_tab_index = self.scenario_tabs.currentIndex()
        if current_tab_index not in self.tab_grid_occupancy:
            return

        current_pos = api_block_widget.property("grid_pos")
        if not current_pos:
            return

        row, col = current_pos
        target_row, target_col = row, col

        # Determine the target position
        if row > 0: # Move up within the same column
            target_row = row - 1
            target_col = col
        elif col > 0: # Move to the last position of the previous column
            target_row = 7 # Last row
            target_col = col - 1
        else: # Already at (0,0), cannot move up
            show_message("Move Failed","Cannot move up: Already at the top-leftmost position.",level="info")
            return

        # Check if the target position is valid (within grid bounds)
        if not (0 <= target_row < 8 and 0 <= target_col < 4):
            # This case should theoretically be covered by the initial if/elif/else,
            # but acts as a safeguard.
            show_message("Move Failed","Cannot move up: Target position out of bounds.",level="info")
            return

        target_widget = self.current_tab_work_layout.itemAtPosition(target_row, target_col)
        
        # Rule 3: If target position is empty, do not allow traversal/movement.
        if (target_row, target_col) not in self.tab_grid_occupancy[current_tab_index]:
            show_message("Move Failed","Cannot move up: Target position is empty.",level="info")
            return

        # Rule 1 & 2: If there's an API block, swap.
        # target_widget should not be None here due to the above check.
        if target_widget and isinstance(target_widget.widget(), ApiBlock):
            # Get the widget at the target position
            other_api_block_widget = target_widget.widget()

            # Remove both widgets from their current positions
            self.current_tab_work_layout.removeWidget(api_block_widget)
            self.current_tab_work_layout.removeWidget(other_api_block_widget)

            # Update occupancy for both
            self.tab_grid_occupancy[current_tab_index].remove(current_pos)
            self.tab_grid_occupancy[current_tab_index].remove((target_row, target_col))

            # Add widgets to their new positions
            self.current_tab_work_layout.addWidget(api_block_widget, target_row, target_col)
            self.current_tab_work_layout.addWidget(other_api_block_widget, row, col)

            # Update grid_pos property for both widgets
            api_block_widget.setProperty("grid_pos", (target_row, target_col))
            other_api_block_widget.setProperty("grid_pos", (row, col))

            self.tab_grid_occupancy[current_tab_index].add((target_row, target_col))
            self.tab_grid_occupancy[current_tab_index].add((row, col))

            print(f"Swapped API block '{api_block_widget.api_name}' from {current_pos} to {(target_row, target_col)} "
                  f"with '{other_api_block_widget.api_name}'.")
            
            # Ensure layout updates
            self.current_tab_work_area_content_widget.updateGeometry()
            self.current_tab_work_area_content_widget.repaint()


    def _move_api_block_down(self, api_block_widget: QWidget):
        self.on_change = True
        if not self.current_tab_work_layout:
            return

        current_tab_index = self.scenario_tabs.currentIndex()
        if current_tab_index not in self.tab_grid_occupancy:
            return

        current_pos = api_block_widget.property("grid_pos")
        if not current_pos:
            return

        row, col = current_pos
        target_row, target_col = row, col

        # Determine the target position
        if row < 7: # Move down within the same column
            target_row = row + 1
            target_col = col
        elif col < 3: # Move to the first position of the next column
            target_row = 0 # First row
            target_col = col + 1
        else: # Already at (7,3), cannot move down
            show_message("Move Failed","Cannot move down: Already at the bottom-rightmost position.",level="info")
            return

        # Check if the target position is valid (within grid bounds)
        if not (0 <= target_row < 8 and 0 <= target_col < 4):
            # This case should theoretically be covered by the initial if/elif/else,
            # but acts as a safeguard.
            show_message("Move Failed","Cannot move down: Target position out of bounds.",level="info")
            return
            
        target_widget = self.current_tab_work_layout.itemAtPosition(target_row, target_col)

        # Rule 3: If target position is empty, do not allow traversal/movement.
        if (target_row, target_col) not in self.tab_grid_occupancy[current_tab_index]:
            show_message("Move Failed","Cannot move down: Target position is empty.",level="info")
            return

        # Rule 1 & 2: If there's an API block, swap.
        # target_widget should not be None here due to the above check.
        if target_widget and isinstance(target_widget.widget(), ApiBlock):
            # Get the widget at the target position
            other_api_block_widget = target_widget.widget()

            # Remove both widgets from their current positions
            self.current_tab_work_layout.removeWidget(api_block_widget)
            self.current_tab_work_layout.removeWidget(other_api_block_widget)

            # Update occupancy for both
            self.tab_grid_occupancy[current_tab_index].remove(current_pos)
            self.tab_grid_occupancy[current_tab_index].remove((target_row, target_col))

            # Add widgets to their new positions
            self.current_tab_work_layout.addWidget(api_block_widget, target_row, target_col)
            self.current_tab_work_layout.addWidget(other_api_block_widget, row, col)

            # Update grid_pos property for both widgets
            api_block_widget.setProperty("grid_pos", (target_row, target_col))
            other_api_block_widget.setProperty("grid_pos", (row, col))

            self.tab_grid_occupancy[current_tab_index].add((target_row, target_col))
            self.tab_grid_occupancy[current_tab_index].add((row, col))

            print(f"Swapped API block '{api_block_widget.api_name}' from {current_pos} to {(target_row, target_col)} "
                  f"with '{other_api_block_widget.api_name}'.")

            # Ensure layout updates
            self.current_tab_work_area_content_widget.updateGeometry()
            self.current_tab_work_area_content_widget.repaint()


    def save_project_data(self, file_path):
        """
        Collects all scenario data and saves it to the specified JSON file.
        """
        if not self.scenario_tabs or self.scenario_tabs.count() == 0:
            show_message("Save Project","No scenarios to save.",level="info")
            return False

        project_data = {"scenarios": {}}

        for i in range(self.scenario_tabs.count()):
            tab_name = self.scenario_tabs.tabText(i)
            tab_widget = self.scenario_tabs.widget(i)
            
            # Find the QGridLayout within this tab
            # This requires navigating the layout structure specific to add_scenario_work_section
            api_names_in_scenario = []
            
            # Assuming the structure: QHBoxLayout -> QFrame (right side) -> QVBoxLayout -> QScrollArea -> QWidget (content) -> QGridLayout
            page_main_layout = tab_widget.layout()
            if page_main_layout and page_main_layout.count() > 1:
                work_area_container = page_main_layout.itemAt(1).widget()
                if isinstance(work_area_container, QFrame):
                    container_layout = work_area_container.layout()
                    if container_layout and container_layout.count() > 0:
                        work_scroll_area = container_layout.itemAt(0).widget()
                        if isinstance(work_scroll_area, QScrollArea):
                            content_widget = work_scroll_area.widget()
                            if isinstance(content_widget.layout(), QGridLayout):
                                grid_layout = content_widget.layout()

                                # Collect API blocks in column-wise, then row-wise order
                                # Iterate through columns first, then rows, to get sequential order
                                for col in range(grid_layout.columnCount()):
                                    for row in range(grid_layout.rowCount()):
                                        item = grid_layout.itemAtPosition(row, col)
                                        if item is not None:
                                            widget = item.widget()
                                            if isinstance(widget, ApiBlock):
                                                api_names_in_scenario.append(widget.api_name)
            
            project_data["scenarios"][tab_name] = api_names_in_scenario

        try:
            # Create the 'configs' directory if it doesn't exist
            os.makedirs("projects", exist_ok=True)
            full_path = os.path.join("projects", os.path.basename(file_path))

            with open(full_path, 'w') as f:
                json.dump(project_data, f, indent=4)
            self.current_project_file = full_path # Update the current project file
            show_message( "Save Project",f"Project saved successfully to {full_path}",level="info")
            return True
        except Exception as e:
            show_message("Save Error",f"Failed to save project: {str(e)}",level="critical")
            return False

    def save_project(self):
        self.on_change = False
        """
        Saves the project to the current project file path.
        If no path is set, calls save_project_as.
        """
        if self.current_project_file:
            self.save_project_data(self.current_project_file)
        else:
            self.save_project_as()

    def save_project_as(self):
        self.on_change = False
        """
        Prompts the user for a new file path and saves the project there.
        """
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Project As", 
            os.path.join("projects", "my_project.json"), # Default path and filename
            "JSON Files (*.json)"
        )
        if file_path:
            self.save_project_data(file_path)

    def get_api_blocks(self):
        """
        Fetches all ApiBlock widgets from the QGridLayout and returns them as a list.
        
        Returns:
            list: A list of ApiBlock instances currently in the layout.
        """
        api_blocks_list = []
        
        # Iterate through all items in the layout
        # The .count() method returns the total number of items (widgets or other layouts)
        for i in range(self.current_tab_work_layout.count()):
            # Get the layout item at the current index
            item = self.current_tab_work_layout.itemAt(i)
            
            # Check if the item is a widget and not a spacer or another layout
            if item.widget() is not None:
                widget = item.widget()
                
                # Check if the widget is an instance of the ApiBlock class
                if isinstance(widget, ApiBlock):
                    api_blocks_list.append(widget)
                    
        return api_blocks_list


    def launch_interaction_window(self):

        # Find the QGridLayout within this tab
        # This requires navigating the layout structure specific to add_scenario_work_section
        api_names_in_scenario = []
        active_scenario_widget = self.scenario_tabs.currentWidget()

        # Assuming the structure: QHBoxLayout -> QFrame (right side) -> QVBoxLayout -> QScrollArea -> QWidget (content) -> QGridLayout
        page_main_layout = active_scenario_widget.layout()
        if page_main_layout and page_main_layout.count() > 1:
            work_area_container = page_main_layout.itemAt(1).widget()
            if isinstance(work_area_container, QFrame):
                container_layout = work_area_container.layout()
                if container_layout and container_layout.count() > 0:
                    work_scroll_area = container_layout.itemAt(0).widget()
                    if isinstance(work_scroll_area, QScrollArea):
                        content_widget = work_scroll_area.widget()
                        if isinstance(content_widget.layout(), QGridLayout):
                            grid_layout = content_widget.layout()

                            # Collect API blocks in column-wise, then row-wise order
                            # Iterate through columns first, then rows, to get sequential order
                            for col in range(grid_layout.columnCount()):
                                for row in range(grid_layout.rowCount()):
                                    item = grid_layout.itemAtPosition(row, col)
                                    if item is not None:
                                        widget = item.widget()
                                        if isinstance(widget, ApiBlock):
                                            api_names_in_scenario.append(widget.api_name)
        
        # 2. Create an instance of the new window, passing the API names
        self.interaction_window = MainWindow(api_names_in_scenario, self.scenario_name)
        # 3. Show the new window
        self.interaction_window.show()

    def on_run_backend(self):
        # Call your backend when button is clicked
        self.save_project()
        startEngine.runBackend(self.scenario_name, self.get_api_blocks(), self.current_project_file)
    
    def closeEvent(self, event):
        if self.current_tab_work_area_content_widget != None and self.on_change == True:
            reply = show_message("Confirmation", 'Are you sure you want to exit?', level="question")

            if reply == QMessageBox.Yes:  # Yes clicked
                event.accept()
            else:      # No clicked
                event.ignore()
        else:
            event.accept()


if __name__ == '__main__':
    from PyQt5.QtCore import QT_VERSION_STR
    print(f"Running with Qt Version: {QT_VERSION_STR}")

    app = QApplication(sys.argv)
    window = FlowTestStudio()
    window.show()
    sys.exit(app.exec_())

